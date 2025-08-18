from openpyxl import Workbook
from openpyxl import load_workbook
import itertools

year='m1'
wb = load_workbook(f'stages-{year}.xlsx',)
print('sheets:',wb.sheetnames)


ws = wb['exportConvention']
names=ws['C']
firstnames = ws['D']
codes = ws['L']
sujets = ws['T']
entreprises = ws['BC']
www_entreprises = ws['BR']

masterlist = (('m1', 'MI6251'), ('m2', 'MI6252'))

def writeRapports(f, n, fn, s, e, w ):
    f.write(
"""
 - [[[{0}]]] {1} {2}, _{4}_, link:{5}[{3}], xref:attachment${1}-{2}.pdf[{1}-{2}.pdf],  xref:attachment${1}-{2}-slides.pdf[{1}-{2}-slides.pdf] 
""".format(n.value.title().replace(" ", ""), n.value.title(), fn.value.title(), e.value.title().strip(), s.value.capitalize().strip(), w.value.strip()))


def writeTableEntry(f, n, fn, s, e, w):
    f.write(
        """
| {0} | {1} | link:{4}[{2}] | _{3}_
""".format(n.value.title(), fn.value.title(), e.value.title().strip(), s.value.strip(), w.value.strip()))


for module,master in masterlist:
    if module == year:
        f = open("modules/"+module+"/partials/rapports.adoc", "w")
        for n, fn, c, s, e,w in sorted(zip(names, firstnames, codes, sujets, entreprises,www_entreprises), key=lambda x: x[0].value):
            if c.value == master :
                writeRapports(f, n, fn, s, e, w)
        f.close()
        f = open("modules/"+module+"/partials/stages.adoc", "w")
        f.write('[cols="1,1,2,4"]\n|===\n')
        f.write('| Nom | Prénom | Entreprise | Sujet\n')
        for n, fn, c, s, e,w in sorted(zip(names, firstnames, codes, sujets, entreprises,www_entreprises), key=lambda x: x[0].value):
            if c.value == master :
                writeTableEntry(f, n, fn, s, e, w)
        f.write('\n|===')
        f.close()

encadrants = ws['CB']

emails={'m1':' ','m2':' '}
for module, master in masterlist:
    for n,c, e in zip(names,codes,encadrants):
        if c.value == master:
            emails[module]+=e.value.strip()+','

def writeEmails(master, emails):
    email_string = emails[master]
    email_list = email_string.split(',')
    # Remove any empty strings in case there's a trailing comma
    email_list = [email for email in email_list if email]

    # Convert the list to a set to remove duplicates, then back to a sorted list
    unique_emails_sorted = sorted(set(email_list))

    # Join the unique, sorted emails into a single string separated by commas
    unique_email_string = ', '.join(unique_emails_sorted)
    print(master,unique_email_string)

writeEmails('m1', emails)    
writeEmails('m2', emails)
