from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as colidx

# ---------- config ----------
YEAR = "m2"
XLSX = "stages-2025.xlsx"
SHEET = "exportConvention"

# Column letters (adjust if needed)
COLS = {
    "name": "C",
    "firstname": "D",
    "code": "L",
    "subject": "T",
    "company": "BC",
    "company_url": "BR",
    "advisor_email": "CB",
}

MASTERLIST = (("m1", "MI6251"), ("m2", "MI6252"))

# ---------- helpers ----------
def s(x, default=""):
    """Safe string: None -> '', strip whitespace."""
    return (str(x).strip() if x is not None else default)

def namecase(x):
    """Simple name casing without overzealous .title()."""
    t = s(x)
    # Lowercase everything then capitalize tokens; preserve apostrophes/hyphens
    out = []
    for part in t.split():
        tokens = []
        for tok in part.replace("-", " - ").replace("'", " ' ").split():
            tokens.append(tok[:1].upper() + tok[1:].lower())
        out.append("".join(tokens).replace(" - ", "-").replace(" ' ", "'"))
    return " ".join(out)

def anchor_from_name(x):
    """AsciiDoc anchor without spaces/punctuations."""
    return "".join(ch for ch in s(x).title() if ch.isalnum())

def write_rapports_line(f, rec):
    # rec fields already sanitized
    anchor = anchor_from_name(rec["name"])
    lastname = namecase(rec["name"])
    firstname = namecase(rec["firstname"])
    subject = s(rec["subject"]).capitalize()
    company = namecase(rec["company"])
    url = s(rec["company_url"])

    # If company URL is missing, show plain text; otherwise make link
    company_disp = f"link:{url}[{company}]" if url else company

    f.write(
        f"""
 - [[[{anchor}]]] {lastname} {firstname}, _{company}_, {company_disp if url else company}, \
xref:attachment${lastname}-{firstname}.pdf[{lastname}-{firstname}.pdf], \
xref:attachment${lastname}-{firstname}-slides.pdf[{lastname}-{firstname}-slides.pdf]
"""
    )

def write_table_row(f, rec):
    lastname = namecase(rec["name"])
    firstname = namecase(rec["firstname"])
    subject = s(rec["subject"])
    company = namecase(rec["company"])
    url = s(rec["company_url"])
    company_disp = f"link:{url}[{company}]" if url else company
    f.write(f"\n| {lastname} | {firstname} | {company_disp} | _{subject}_")

# ---------- load workbook ----------
wb = load_workbook(XLSX, data_only=True, read_only=True)
ws = wb[SHEET]

# Map to column indices (1-based)
idx = {k: colidx(v) for k, v in COLS.items()}

# Build a list of records; skip header (min_row=2)
records = []
for row in ws.iter_rows(min_row=2, values_only=True):
    rec = {
        "name": row[idx["name"] - 1],
        "firstname": row[idx["firstname"] - 1],
        "code": row[idx["code"] - 1],
        "subject": row[idx["subject"] - 1],
        "company": row[idx["company"] - 1],
        "company_url": row[idx["company_url"] - 1],
        "advisor_email": row[idx["advisor_email"] - 1],
    }
    # Skip completely empty rows
    if not any(rec.values()):
        continue
    records.append(rec)

# ---------- per-module generation ----------
base = Path("modules")
emails = {"m1": set(), "m2": set()}

for module, master in MASTERLIST:
    # Filter & sort safely
    subset = [
        r for r in records if s(r["code"]) == master
    ]
    subset.sort(key=lambda r: s(r["name"]).casefold())

    if module == YEAR and subset:
        outdir = base / module / "partials"
        outdir.mkdir(parents=True, exist_ok=True)

        # rapports.adoc
        with (outdir / "rapports.adoc").open("w", encoding="utf-8") as f:
            for rec in subset:
                write_rapports_line(f, rec)

        # stages.adoc
        with (outdir / "stages.adoc").open("w", encoding="utf-8") as f:
            f.write('[cols="1,1,2,4"]\n|===\n')
            f.write("| Nom | Prénom | Entreprise | Sujet\n")
            for rec in subset:
                write_table_row(f, rec)
            f.write("\n|===\n")

    # collect emails (unique)
    for r in subset:
        em = s(r["advisor_email"])
        if em:
            # handle possible "Name <email>" or multiple separated by ; or ,
            for token in em.replace(";", ",").split(","):
                token = token.strip()
                if token:
                    emails[module].add(token)

# ---------- print email lists ----------
def print_emails(mod):
    sorted_unique = sorted(emails.get(mod, []), key=str.casefold)
    print(mod, ", ".join(sorted_unique))

print_emails("m1")
print_emails("m2")